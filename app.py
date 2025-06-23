
import streamlit as st
import pandas as pd
import re
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

def normalize_name(name):
    if pd.isnull(name):
        return ""
    name = name.strip().replace("ه", "ة").replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    name = re.sub(r'(عبد)([^\s])', r'\1 \2', name)
    return " ".join(name.split()).lower()

def is_first_three_words_match(name1, name2):
    words1 = name1.split()
    words2 = name2.split()
    length = min(len(words1), len(words2), 3)
    return all(words1[i] == words2[i] for i in range(length))

def match_names(names_df, database_df, mode='name'):
    names_df["normalized_name"] = names_df["اسم الموظف"].apply(normalize_name)
    database_df["normalized_name"] = database_df["اسم الموظف"].apply(normalize_name)

    if mode == 'name':
        database_df = database_df.drop_duplicates(subset=["normalized_name"])
        database_map = database_df.set_index("normalized_name")[["اسم الموظف", "Iban"]].to_dict(orient="index")
    else:
        database_df = database_df.drop_duplicates(subset=["normalized_name"])
        database_map = database_df.set_index("normalized_name")[["اسم الموظف", "Operator Id", "المدرسة", "الدائرة"]].to_dict(orient="index")

    matched_results = []

    for original_name, normalized_name in zip(names_df["اسم الموظف"], names_df["normalized_name"]):
        best_match = None
        best_score = 0
        for db_name in database_map.keys():
            score = fuzz.ratio(normalized_name, db_name)
            if score > best_score:
                best_score = score
                best_match = db_name

        match_data = None
        if best_score >= 85 and (is_first_three_words_match(normalized_name, best_match) or best_match.startswith(normalized_name)):
            match_data = database_map[best_match]
        else:
            for db_name in database_map.keys():
                if db_name.startswith(normalized_name):
                    match_data = database_map[db_name]
                    best_match = db_name
                    best_score = fuzz.ratio(normalized_name, best_match)
                    break

        if match_data:
            result = {
                "الاسم الأصلي": original_name,
                "الاسم المطابق": match_data["اسم الموظف"],
                "نسبة التطابق": f"{round(best_score)}%",
                "ملاحظة": "✅ تطابق دقيق"
            }
            if mode == 'name':
                result["الآيبان"] = match_data["Iban"]
            else:
                result["Operator Id"] = match_data.get("Operator Id", "")
                result["المدرسة"] = match_data.get("المدرسة", "")
                result["الدائرة"] = match_data.get("الدائرة", "")
            matched_results.append(result)
        else:
            result = {
                "الاسم الأصلي": original_name,
                "الاسم المطابق": "",
                "نسبة التطابق": "",
                "ملاحظة": "❌ لم يتم العثور على تطابق"
            }
            if mode == 'name':
                result["الآيبان"] = ""
            else:
                result["Operator Id"] = ""
                result["المدرسة"] = ""
                result["الدائرة"] = ""
            matched_results.append(result)

    return pd.DataFrame(matched_results)

def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        wb = writer.book
        ws = wb.active

        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            col_letter = column[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            note = row[3].value
            if note and "❌" in note:
                for cell in row:
                    cell.fill = red_fill

    output.seek(0)
    return output

# ========== Streamlit UI ==========
st.set_page_config(page_title="تطبيق المطابقة", layout="centered")
st.title("🔐محمد حماية الدخول")

password = st.text_input("أدخل كلمة المرور:", type="password")

if password == "mjaleel":
    st.success("✅ تم التحقق من كلمة المرور.")
    tab1, tab2 = st.tabs(["مطابقة الأسماء", "مطابقة الأقسام"])

    with tab1:
        st.subheader("رفع ملفات مطابقة الأسماء")
        file1 = st.file_uploader("📄 ملف الأسماء", type="xlsx", key="file1_name")
        file2 = st.file_uploader("📊 ملف قاعدة البيانات", type="xlsx", key="file2_name")

        if file1 and file2:
            if st.button("🚀 بدء المطابقة"):
                names_df = pd.read_excel(file1)
                db_df = pd.read_excel(file2)
                results = match_names(names_df, db_df, mode='name')
                st.success("✅ تم المطابقة.")
                st.dataframe(results)
                excel_data = to_excel(results)
                st.download_button("⬇️ تحميل الملف النهائي", excel_data, file_name="نتائج_المطابقة.xlsx")

    with tab2:
        st.subheader("رفع ملفات مطابقة الأقسام")
        file3 = st.file_uploader("📄 ملف الأسماء", type="xlsx", key="file1_dept")
        file4 = st.file_uploader("📊 ملف قاعدة البيانات", type="xlsx", key="file2_dept")

        if file3 and file4:
            if st.button("🚀 بدء المطابقة", key="match_dept"):
                names_df = pd.read_excel(file3)
                db_df = pd.read_excel(file4)
                results = match_names(names_df, db_df, mode='department')
                st.success("✅ تم المطابقة.")
                st.dataframe(results)
                excel_data = to_excel(results)
                st.download_button("⬇️ تحميل الملف النهائي", excel_data, file_name="نتائج_الأقسام.xlsx")

elif password:
    st.error("❌ كلمة المرور غير صحيحة.")
